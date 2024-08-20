!pip install PyPDF2

!pip install azure-core

!pip install azure-ai-formrecognizer

!pip install openpyxl







from PyPDF2 import PdfWriter, PdfReader

import openpyxl

from azure.core.credentials import AzureKeyCredential

from azure.ai.formrecognizer import DocumentAnalysisClient







# Split the PDF into multiple pages

inputpdf = PdfReader(open("example.pdf", "rb"))

workbook = openpyxl.Workbook()

 

# Initialize the Form Recognizer client

endpoint = "your api-endpoint"

key = "your api-key"

document_analysis_client = DocumentAnalysisClient(

    endpoint=endpoint, credential=AzureKeyCredential(key)

)

 

for i in range(len(inputpdf.pages)):

    output = PdfWriter()

    output.add_page(inputpdf.pages[i])

    with open("your.pdf" % i, "wb") as outputStream:

        output.write(outputStream)

 

    # Analyze the page using Form Recognizer

    with open("/content/sample_data/123/document-page%s.pdf" % i, "rb") as pdf_file:

      #make a folder of 123 for pages

        poller = document_analysis_client.begin_analyze_document("prebuilt-layout", pdf_file)

        result = poller.result()

        tables = result.tables

 

        # Check if there are tables in the page

        if tables:

            # Create a new sheet for each page

            sheet = workbook.create_sheet(f"Page {i+1}")

 

            # Initialize row offset

            row_offset = 1

 

            # Iterate through each table on the page

            for table_index, table in enumerate(tables):

                # Initialize an empty list to store rows

                rows = [['' for _ in range(table.column_count)] for _ in range(table.row_count)]

 

                for cell in table.cells:

                    for row in range(cell.row_index, cell.row_index + cell.row_span):

                        for col in range(cell.column_index, cell.column_index + cell.column_span):

                            rows[row][col] = cell.content

 

                # Write the rows to the Excel sheet with offset

                for row_index, row in enumerate(rows, start=row_offset):

                    for col_index, value in enumerate(row, start=1):

                        cell = sheet.cell(row=row_index, column=col_index)

                        cell.value = value

 

                # Increment row offset for the next table

                row_offset += table.row_count + 2  # Add 2 for some spacing between tables

 

# Save the Excel workbook

workbook.save('book.xlsx')